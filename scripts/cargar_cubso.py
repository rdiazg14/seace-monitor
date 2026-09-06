#!/usr/bin/env python3
"""C5: carga el XLS/XLSX del CUBSO (OECE) a cubso_catalogo.

El archivo publicado es XLSX con extension .xls (magic PK). Datos desde la
fila 7. No toca it_keywords.

    uv run python scripts/cargar_cubso.py data/cubso/cubso_oece.xls --version 2016-03-28
    uv run python scripts/cargar_cubso.py data/cubso/cubso_oece.xls --version 2016-03-28 --dry-run
"""
from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
import tempfile
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import psycopg
from openpyxl import load_workbook
from psycopg.rows import dict_row

_ROOT = Path(__file__).resolve().parent.parent
_ENV = _ROOT / ".env"

FUENTE_DEFAULT = (
    "https://www.gob.pe/institucion/oece/informes-publicaciones/"
    "5813164-anexo-catalogo-unico-de-bienes-servicios-y-obras-cubso"
)

TIPOS_OK = {
    "BIENES": "BIENES",
    "SERVICIOS": "SERVICIOS",
    "OBRAS": "OBRAS",
    "CONSULTORIAS OBRAS": "CONSULTORIAS OBRAS",
    "CONSULTORIA OBRAS": "CONSULTORIAS OBRAS",
    "CONSULTORIAS DE OBRAS": "CONSULTORIAS OBRAS",
}

CODIGO_RE = re.compile(r"^\d{16}$")
BATCH = 5000


def _cargar_env() -> None:
    if not _ENV.exists():
        return
    for line in _ENV.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip())


def _dsn() -> str:
    _cargar_env()
    dsn = (os.environ.get("DATABASE_URL") or "").strip()
    if not dsn:
        print("ERROR: falta DATABASE_URL en el entorno o en .env", flush=True)
        sys.exit(2)
    return dsn


def _redactar(texto: str) -> str:
    url = os.environ.get("DATABASE_URL") or ""
    if url:
        texto = texto.replace(url, "***")
    return re.sub(
        r"postgres(?:ql)?(?:\+[a-z]+)?://[^\s'\"<>]+", "***", texto, flags=re.I
    )


def _ruta_xlsx(origen: Path) -> tuple[Path, Path | None]:
    """openpyxl rechaza .xls aunque el magic sea PK (xlsx). Copia a temp si hace falta."""
    head = origen.read_bytes()[:8]
    if head[:2] != b"PK":
        print(
            f"ERROR: {origen} no es XLSX (magic {head[:4]!r}). "
            "El CUBSO del OECE es spreadsheetml con extension .xls.",
            flush=True,
        )
        sys.exit(1)
    if origen.suffix.lower() == ".xlsx":
        return origen, None
    fd, name = tempfile.mkstemp(prefix="cubso_", suffix=".xlsx")
    os.close(fd)
    tmp = Path(name)
    shutil.copyfile(origen, tmp)
    return tmp, tmp


def _codigo(raw) -> str | None:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    if isinstance(raw, int):
        s = f"{raw:016d}"
    elif isinstance(raw, float):
        s = f"{raw:.0f}"
    else:
        s = str(raw).strip()
        if s.endswith(".0"):
            s = s[:-2]
        s = re.sub(r"\s+", "", s)
    if not CODIGO_RE.match(s):
        return None
    return s


def _tipo(raw) -> str | None:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    t = str(raw).strip().upper()
    t = (
        t.replace("Á", "A")
        .replace("É", "E")
        .replace("Í", "I")
        .replace("Ó", "O")
        .replace("Ú", "U")
    )
    if "-" in t:
        t = t.split("-", 1)[1].strip()
    return TIPOS_OK.get(t)


def leer_filas(ruta: Path) -> tuple[list[dict], list[dict], int]:
    """Datos desde la fila 7. pandas+openpyxl; codigo se lee celda a celda (no float)."""
    xlsx, tmp = _ruta_xlsx(ruta)
    try:
        # dtype=str no basta: un float IEEE redondea 16 digitos. openpyxl da int de Python.
        wb = load_workbook(xlsx, read_only=True, data_only=True)
        ws = wb.active
        records: list[dict] = []
        for excel_row, row in enumerate(
            ws.iter_rows(min_row=7, max_col=4, values_only=True), start=7
        ):
            nro, codigo, titulo, tipo = (list(row) + [None, None, None, None])[:4]
            if codigo is None and titulo is None and tipo is None:
                continue
            records.append({
                "excel_row": excel_row,
                "nro": nro,
                "codigo": codigo,
                "titulo": titulo,
                "tipo": tipo,
            })
        wb.close()
        df = pd.DataFrame.from_records(records)
    finally:
        if tmp is not None:
            tmp.unlink(missing_ok=True)

    n_leidos = 0
    ok: list[dict] = []
    desc: list[dict] = []
    vistos: set[str] = set()
    for rec in df.to_dict("records"):
        n_leidos += 1
        excel_row = int(rec["excel_row"])
        codigo = _codigo(rec["codigo"])
        tipo = _tipo(rec["tipo"])
        titulo_raw = rec["titulo"]
        titulo = "" if titulo_raw is None or (
            isinstance(titulo_raw, float) and pd.isna(titulo_raw)
        ) else str(titulo_raw).strip()
        if not codigo:
            desc.append({
                "fila": excel_row,
                "motivo": "codigo_invalido",
                "codigo": None if rec["codigo"] is None else str(rec["codigo"]),
            })
            continue
        if not titulo:
            desc.append({"fila": excel_row, "motivo": "titulo_vacio", "codigo": codigo})
            continue
        if not tipo:
            desc.append({
                "fila": excel_row,
                "motivo": "tipo_invalido",
                "codigo": codigo,
                "tipo": None if rec["tipo"] is None else str(rec["tipo"]),
            })
            continue
        if codigo in vistos:
            desc.append({"fila": excel_row, "motivo": "duplicado_archivo", "codigo": codigo})
            continue
        vistos.add(codigo)
        ok.append({
            "codigo": codigo,
            "titulo": titulo,
            "tipo": tipo,
            "segmento": codigo[:2],
            "familia": codigo[:4],
            "clase": codigo[:6],
            "commodity": codigo[:8],
        })
    return ok, desc, n_leidos


def _upsert(conn, filas: list[dict], version: date) -> tuple[int, int]:
    insertados = 0
    actualizados = 0
    sql = """
        INSERT INTO cubso_catalogo (
            codigo, titulo, tipo, segmento, familia, clase, commodity, version_catalogo
        )
        SELECT
            x.codigo, x.titulo, x.tipo, x.segmento, x.familia, x.clase,
            x.commodity, %s
        FROM jsonb_to_recordset(%s::jsonb) AS x(
            codigo text, titulo text, tipo text,
            segmento text, familia text, clase text, commodity text
        )
        ON CONFLICT (codigo) DO UPDATE SET
            titulo = EXCLUDED.titulo,
            tipo = EXCLUDED.tipo,
            segmento = EXCLUDED.segmento,
            familia = EXCLUDED.familia,
            clase = EXCLUDED.clase,
            commodity = EXCLUDED.commodity,
            version_catalogo = EXCLUDED.version_catalogo,
            cargado_utc = now()
        RETURNING (xmax = 0) AS inserted
    """
    for i in range(0, len(filas), BATCH):
        chunk = filas[i:i + BATCH]
        payload = json.dumps(chunk, ensure_ascii=False)
        rows = conn.execute(sql, (version, payload)).fetchall()
        for r in rows:
            if r["inserted"]:
                insertados += 1
            else:
                actualizados += 1
        print(
            f"  lote {i // BATCH + 1}: {len(chunk)} "
            f"(acum ins={insertados} upd={actualizados})",
            flush=True,
        )
    return insertados, actualizados


def main() -> int:
    ap = argparse.ArgumentParser(description="Carga CUBSO OECE a cubso_catalogo")
    ap.add_argument("ruta_xls", help="XLS/XLSX del catalogo (fila 7 = primer item)")
    ap.add_argument(
        "--version",
        required=True,
        help="Fecha de publicacion del archivo (YYYY-MM-DD)",
    )
    ap.add_argument(
        "--fuente-url",
        default=FUENTE_DEFAULT,
        help="URL de la ficha gob.pe/oece (queda en cubso_version)",
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Valida y reporta sin escribir",
    )
    args = ap.parse_args()

    ruta = Path(args.ruta_xls)
    if not ruta.is_file():
        print(f"ERROR: no existe {ruta}", flush=True)
        return 1
    try:
        version = date.fromisoformat(args.version)
    except ValueError:
        print("ERROR: --version debe ser YYYY-MM-DD", flush=True)
        return 1

    print(f"leyendo {ruta} ({ruta.stat().st_size} bytes) ...", flush=True)
    ok, desc, n_leidos = leer_filas(ruta)
    por_tipo = Counter(r["tipo"] for r in ok)
    por_seg = Counter(r["segmento"] for r in ok)
    motivos = Counter(d["motivo"] for d in desc)

    print(f"leidos (filas con dato): {n_leidos}", flush=True)
    print(f"validos: {len(ok)}", flush=True)
    print(f"descartados: {len(desc)} {dict(motivos)}", flush=True)
    print("tipo:", dict(por_tipo), flush=True)
    print("segmentos distintos:", len(por_seg), flush=True)
    print("top segmentos:", por_seg.most_common(12), flush=True)
    if desc[:8]:
        print("muestra descartados:", desc[:8], flush=True)

    if args.dry_run:
        print("DRY-RUN: no se escribio nada.", flush=True)
        return 0

    dsn = _dsn()
    try:
        with psycopg.connect(dsn, row_factory=dict_row) as conn:
            ins, upd = _upsert(conn, ok, version)
            conn.execute(
                """
                INSERT INTO cubso_version (
                    id, version_catalogo, fuente_url, items, cargado_utc
                )
                VALUES (1, %s, %s, %s, now())
                ON CONFLICT (id) DO UPDATE SET
                    version_catalogo = EXCLUDED.version_catalogo,
                    fuente_url = EXCLUDED.fuente_url,
                    items = EXCLUDED.items,
                    cargado_utc = now()
                """,
                (version, args.fuente_url, len(ok)),
            )
            conn.commit()
    except Exception as e:
        print(_redactar(f"{type(e).__name__}: {e}"), file=sys.stderr, flush=True)
        return 1

    print(
        f"OK insertados={ins} actualizados={upd} descartados={len(desc)} "
        f"leidos={n_leidos} version={version.isoformat()} "
        f"cargado_utc={datetime.now().astimezone().isoformat(timespec='seconds')}",
        flush=True,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
